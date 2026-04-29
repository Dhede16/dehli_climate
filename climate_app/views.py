"""
Delhi Climate — views.py
Satu metode imputasi: Cubic Spline Interpolation.
Endpoint:
  GET  /             → halaman utama
  GET  /api/data/    → data mentah JSON
  POST /api/impute/  → imputasi semua kolom, return JSON
  GET  /api/download/→ download Excel hasil imputasi
"""

import io
import math

import pandas as pd
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt

# ── Konfigurasi ────────────────────────────────────────────────────────────────

COLUMNS = ['meantemp', 'humidity', 'wind_speed', 'meanpressure']

ALGO_NAME = 'Cubic Spline Interpolation'
ALGO_DESC = (
    'Mengisi nilai kosong menggunakan kurva kubik yang mulus yang '
    'menghubungkan titik-titik data yang diketahui.'
    ' Metode ini mempertimbangkan nilai sebelum '
    'dan sesudah setiap titik kosong sehingga hasil interpolasi lebih '
    'alami dan akurat untuk data iklim deret waktu.'
)


# ── Data ───────────────────────────────────────────────────────────────────────

def load_data():
    """Baca CSV, kembalikan list of dict (NaN → None)."""
    df = pd.read_csv(settings.DATA_FILE)
    rows = df.to_dict(orient='records')
    for row in rows:
        for k, v in row.items():
            if isinstance(v, float) and math.isnan(v):
                row[k] = None
    return rows


# ── Algoritma Cubic Spline ─────────────────────────────────────────────────────

def cubic_spline(arr):
    """
    Interpolasi Cubic Hermite Spline pada array 1-D.
    - Temukan semua titik yang diketahui (bukan None).
    - Untuk setiap celah antar dua anchor, hitung slope tangensial
      menggunakan central-difference (atau one-sided di tepi).
    - Evaluasi basis Hermite untuk setiap indeks kosong.
    - Isi kepala/ekor dengan nilai anchor terdekat.
    """
    n = len(arr)
    known = [{'i': i, 'v': arr[i]} for i in range(n) if arr[i] is not None]

    if not known:
        return list(arr)

    result = list(arr)

    for k in range(len(known) - 1):
        i0, v0 = known[k]['i'],     known[k]['v']
        i1, v1 = known[k + 1]['i'], known[k + 1]['v']
        d = i1 - i0
        if d == 1:
            continue  # titik bersebelahan, tidak ada celah

        # tangent kiri (central-diff jika bukan anchor pertama)
        m0 = ((known[k + 1]['v'] - known[k - 1]['v']) /
              (known[k + 1]['i'] - known[k - 1]['i'])
              if k > 0 else (v1 - v0) / d)

        # tangent kanan (central-diff jika bukan anchor terakhir)
        m1 = ((known[k + 2]['v'] - known[k]['v']) /
              (known[k + 2]['i'] - known[k]['i'])
              if k < len(known) - 2 else (v1 - v0) / d)

        for x in range(i0 + 1, i1):
            t   = (x - i0) / d
            h00 =  2*t**3 - 3*t**2 + 1
            h10 =    t**3 - 2*t**2 + t
            h01 = -2*t**3 + 3*t**2
            h11 =    t**3 -   t**2
            result[x] = h00*v0 + h10*d*m0 + h01*v1 + h11*d*m1

    # padding tepi
    for i in range(known[0]['i']):
        result[i] = known[0]['v']
    for i in range(known[-1]['i'] + 1, n):
        result[i] = known[-1]['v']

    return result


def impute_all(data):
    """
    Terapkan cubic_spline ke semua COLUMNS.
    Return (imputed_data, summary) di mana summary = {col: [{index, date, imputed}]}.
    """
    result  = [dict(row) for row in data]
    summary = {}

    for col in COLUMNS:
        arr     = [r[col] for r in data]
        missing = [i for i, v in enumerate(arr) if v is None]
        filled  = cubic_spline(arr)

        for i in missing:
            result[i][col] = filled[i]

        summary[col] = [
            {'index': i, 'date': data[i]['date'], 'imputed': round(filled[i], 4)}
            for i in missing
        ]

    return result, summary


# ── Views ──────────────────────────────────────────────────────────────────────

def index(request):
    data   = load_data()
    miss   = {col: sum(1 for r in data if r[col] is None) for col in COLUMNS}
    context = {
        'total_rows':      len(data),
        'total_missing':   sum(miss.values()),
        'affected_rows':   sum(1 for r in data if any(r[c] is None for c in COLUMNS)),
        'missing_per_col': miss,
        'algo_name':       ALGO_NAME,
        'algo_desc':       ALGO_DESC,
        'columns':         COLUMNS,
    }
    return render(request, 'climate_app/index.html', context)


def api_data(request):
    """GET /api/data/ — seluruh dataset mentah sebagai JSON."""
    return JsonResponse({'data': load_data()})


@csrf_exempt
def api_impute(request):
    """
    POST /api/impute/
    Tidak perlu request body. Mengimputasi SEMUA kolom sekaligus.
    Response: { data: [...], summary: {col: [...]}, total_filled: N }
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    data, summary = impute_all(load_data())

    return JsonResponse({
        'data':         data,
        'summary':      summary,
        'total_filled': sum(len(v) for v in summary.values()),
    })


def api_download(request):
    """
    GET /api/download/
    Menghasilkan file .xlsx berisi data yang sudah diimputasi.
    Sel yang diisi berwarna hijau muda agar mudah dibedakan.
    """
    raw, summary = impute_all(load_data())

    # kumpulkan koordinat sel yang diimputasi
    imputed_cells = {
        (entry['index'], col)
        for col, entries in summary.items()
        for entry in entries
    }

    # bangun DataFrame
    df = pd.DataFrame(raw)
    df['date'] = pd.to_datetime(df['date'])
    for col in COLUMNS:
        df[col] = df[col].round(4)

    # tulis ke buffer Excel
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Data Lengkap')

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = writer.book
        ws = writer.sheets['Data Lengkap']

        # ── style helpers ──
        FILL_HEADER  = PatternFill('solid', fgColor='1F4E79')
        FILL_IMPUTED = PatternFill('solid', fgColor='C6EFCE')
        FONT_HEADER  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        FONT_IMPUTED = Font(name='Calibri', color='276221', size=10)
        FONT_NORMAL  = Font(name='Calibri', size=10)
        THIN = Side(style='thin', color='D0D0D0')
        BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        CENTER = Alignment(horizontal='center', vertical='center')

        col_map = {
            'date':         ('A', 'Tanggal',              14),
            'meantemp':     ('B', 'Mean Temp (°C)',        16),
            'humidity':     ('C', 'Humidity (%)',          14),
            'wind_speed':   ('D', 'Wind Speed (km/h)',     16),
            'meanpressure': ('E', 'Mean Pressure (hPa)',   18),
        }

        # header row styling
        for col_name, (letter, label, width) in col_map.items():
            cell = ws[f'{letter}1']
            cell.value     = label
            cell.fill      = FILL_HEADER
            cell.font      = FONT_HEADER
            cell.alignment = CENTER
            cell.border    = BORDER
            ws.column_dimensions[letter].width = width

        ws.row_dimensions[1].height = 22
        ws.freeze_panes = 'A2'

        # data rows
        for row_idx in range(len(raw)):
            xl_row = row_idx + 2
            for col_name, (letter, _, _) in col_map.items():
                cell           = ws[f'{letter}{xl_row}']
                cell.alignment = CENTER
                cell.border    = BORDER
                if (row_idx, col_name) in imputed_cells:
                    cell.fill = FILL_IMPUTED
                    cell.font = FONT_IMPUTED
                else:
                    cell.font = FONT_NORMAL

        # ── sheet kedua: ringkasan imputasi ──
        ws2 = wb.create_sheet('Ringkasan Imputasi')
        ws2.column_dimensions['A'].width = 22
        ws2.column_dimensions['B'].width = 14
        ws2.column_dimensions['C'].width = 14
        ws2.column_dimensions['D'].width = 16

        FILL_H2 = PatternFill('solid', fgColor='2E75B6')
        FH2 = Font(name='Calibri', bold=True, color='FFFFFF', size=11)

        # header ringkasan
        for letter, label in zip('ABCD', ['Kolom', 'Jml Hilang', 'Sudah Diisi', 'Sisa Kosong']):
            c = ws2[f'{letter}1']
            c.value = label; c.fill = FILL_H2; c.font = FH2
            c.alignment = CENTER; c.border = BORDER

        for r, col in enumerate(COLUMNS, start=2):
            filled = len(summary[col])
            ws2[f'A{r}'].value = col
            ws2[f'B{r}'].value = filled   # semua sudah diisi
            ws2[f'C{r}'].value = filled
            ws2[f'D{r}'].value = 0
            for letter in 'ABCD':
                c = ws2[f'{letter}{r}']
                c.font = FONT_NORMAL; c.alignment = CENTER; c.border = BORDER

        # total row
        tr = len(COLUMNS) + 2
        total_miss = sum(len(summary[c]) for c in COLUMNS)
        ws2[f'A{tr}'].value = 'TOTAL'
        ws2[f'B{tr}'].value = total_miss
        ws2[f'C{tr}'].value = total_miss
        ws2[f'D{tr}'].value = 0
        for letter in 'ABCD':
            c = ws2[f'{letter}{tr}']
            c.font = Font(name='Calibri', bold=True, size=10)
            c.alignment = CENTER; c.border = BORDER

    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = 'attachment; filename="delhi_climate_imputed.xlsx"'
    return resp